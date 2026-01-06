"""
BUI Attendance Kiosk (Tkinter + Excel)

What you get:
1) Students sheet (registry):
   - Sheet: "Students"
   - Columns: Name | OfficialStatus   (Registered / Unregistered)

2) Attendance matrix sheet (copy-friendly):
   - Sheet: "Attendance"
   - Row 1: FULL NAME | 13-Sep | 20-Sep | ...
   - Cells under date columns:
       - time (e.g., 10:07 AM) = present
       - "A" = absent (filled when you click Finalize Today)
       - blank = not finalized yet / not signed in yet
   - Late highlight: if sign-in time is after 10:15 AM Mountain Time, cell becomes light red.

3) Attendance Log sheet (what you asked for):
   - Sheet: "Attendance Log"
   - Columns:
       Timestamp (MT) | Name | OfficialStatus | Attendance Date | Attendance (P/A)
   - When students sign in: a "P" row is logged with timestamp.
   - When you click Finalize Today: all remaining students get "A" rows (timestamp blank),
     and if any "P" is missing in the log (rare), it will be written using the matrix time.

Key behaviors:
- No duplicate sign-ins for the same date column.
- Admin adding a new student ALWAYS adds them as Unregistered, and they show up in Attendance immediately.
- Finalize Today fills A for everyone who didn't sign in and generates the full P/A log for the day.

Run:
  cd ~/Desktop/BUI_Attendance_Kiosk
  source .venv/bin/activate
  python app/attendance_kiosk.py
"""

import os
from datetime import datetime
import difflib
import tkinter as tk
from tkinter import messagebox
from zoneinfo import ZoneInfo

from openpyxl import Workbook, load_workbook
from openpyxl.utils import get_column_letter
from openpyxl.styles import PatternFill


# -------------------------
# Paths / constants
# -------------------------
PROJECT_ROOT = os.path.dirname(os.path.dirname(__file__))  # .../BUI_Attendance_Kiosk
DATA_DIR = os.path.join(PROJECT_ROOT, "data")
WORKBOOK_PATH = os.path.join(DATA_DIR, "BUI_Attendance.xlsx")

STUDENTS_SHEET = "Students"
ATTENDANCE_SHEET = "Attendance"
LOG_SHEET = "Attendance Log"

STATUS_REGISTERED = "Registered"
STATUS_UNREGISTERED = "Unregistered"

# Mountain Time
MT_TZ = ZoneInfo("America/Edmonton")  # AB Mountain Time (handles DST)
LATE_HOUR = 10
LATE_MINUTE = 15
LATE_FILL = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")  # light red


# -------------------------
# Helpers
# -------------------------
def ensure_data_dir():
    os.makedirs(DATA_DIR, exist_ok=True)


def autosize_columns(ws):
    for col in range(1, ws.max_column + 1):
        max_len = 0
        col_letter = get_column_letter(col)
        for cell in ws[col_letter]:
            if cell.value is not None:
                max_len = max(max_len, len(str(cell.value)))
        ws.column_dimensions[col_letter].width = min(max_len + 2, 45)


def get_or_create_workbook(path):
    if os.path.exists(path):
        return load_workbook(path)
    wb = Workbook()
    wb.save(path)
    return load_workbook(path)


def today_col_label() -> str:
    """
    Column header format: '13-Sep', '4-Oct', etc. (no leading zero on day)
    Uses Mountain Time date.
    """
    now_mt = datetime.now(MT_TZ)
    return f"{now_mt.day}-{now_mt.strftime('%b')}"


# -------------------------
# Students (registry)
# -------------------------
def get_or_create_students_sheet(wb):
    if STUDENTS_SHEET in wb.sheetnames:
        ws = wb[STUDENTS_SHEET]

        # Upgrade old version if needed
        a1 = ws.cell(row=1, column=1).value
        b1 = ws.cell(row=1, column=2).value if ws.max_column >= 2 else None

        if a1 == "Registered Student Name" or (a1 == "Name" and b1 is None):
            ws.cell(row=1, column=1, value="Name")
            ws.cell(row=1, column=2, value="OfficialStatus")
            for r in range(2, ws.max_row + 1):
                if ws.cell(row=r, column=1).value and not ws.cell(row=r, column=2).value:
                    ws.cell(row=r, column=2, value=STATUS_REGISTERED)
            autosize_columns(ws)
            wb.save(WORKBOOK_PATH)

        return ws

    ws = wb.create_sheet(STUDENTS_SHEET)
    ws.append(["Name", "OfficialStatus"])
    autosize_columns(ws)
    wb.save(WORKBOOK_PATH)
    return ws


def load_students(wb):
    """
    Returns:
      students: dict casefold(name) -> (official_name, status)
      names: list of official names (for suggestions)
    """
    ws = get_or_create_students_sheet(wb)
    students = {}
    names = []

    for row in ws.iter_rows(min_row=2, values_only=True):
        if not row or not row[0]:
            continue
        name = str(row[0]).strip()
        status = str(row[1]).strip() if len(row) > 1 and row[1] else STATUS_UNREGISTERED

        key = name.casefold()
        if key not in students:
            students[key] = (name, status)
            names.append(name)

    return students, names


# -------------------------
# Attendance matrix sheet
# -------------------------
def get_or_create_attendance_sheet(wb):
    if ATTENDANCE_SHEET in wb.sheetnames:
        ws = wb[ATTENDANCE_SHEET]
        if ws.max_row < 1 or ws.cell(row=1, column=1).value is None:
            ws.append(["FULL NAME"])
            autosize_columns(ws)
            wb.save(WORKBOOK_PATH)
        return ws

    ws = wb.create_sheet(ATTENDANCE_SHEET)
    ws.append(["FULL NAME"])
    autosize_columns(ws)
    wb.save(WORKBOOK_PATH)
    return ws


def ensure_date_column(ws, date_label: str) -> int:
    for col in range(2, ws.max_column + 1):
        val = ws.cell(row=1, column=col).value
        if val and str(val).strip() == date_label:
            return col

    new_col = ws.max_column + 1
    ws.cell(row=1, column=new_col, value=date_label)
    autosize_columns(ws)
    return new_col


def find_or_create_student_row(ws, name: str) -> int:
    target = name.strip().casefold()

    for r in range(2, ws.max_row + 1):
        val = ws.cell(row=r, column=1).value
        if val and str(val).strip().casefold() == target:
            return r

    new_row = ws.max_row + 1
    ws.cell(row=new_row, column=1, value=name.strip())
    autosize_columns(ws)
    return new_row


def ensure_student_row_in_attendance(wb, name: str):
    ws = get_or_create_attendance_sheet(wb)
    find_or_create_student_row(ws, name)
    autosize_columns(ws)
    wb.save(WORKBOOK_PATH)


# -------------------------
# Attendance log sheet
# -------------------------
def get_or_create_log_sheet(wb):
    if LOG_SHEET in wb.sheetnames:
        ws = wb[LOG_SHEET]
    else:
        ws = wb.create_sheet(LOG_SHEET)

    if ws.max_row < 1 or ws.cell(row=1, column=1).value is None:
        ws.append(["Timestamp (MT)", "Name", "OfficialStatus", "Attendance Date", "Attendance (P/A)"])
        autosize_columns(ws)
        wb.save(WORKBOOK_PATH)

    return ws


# -------------------------
# Add students
# -------------------------
def add_student_as_unregistered(wb, name: str):
    name = name.strip()
    if not name:
        return False, "Name cannot be empty."

    students, _ = load_students(wb)
    if name.casefold() in students:
        return False, "That student already exists in Students."

    ws = get_or_create_students_sheet(wb)
    ws.append([name, STATUS_UNREGISTERED])
    autosize_columns(ws)

    # Ensure they appear in Attendance immediately
    ensure_student_row_in_attendance(wb, name)

    wb.save(WORKBOOK_PATH)
    return True, f"Added as Unregistered: {name}"


# -------------------------
# Matching / suggestions
# -------------------------
def canonical_match(typed: str, students: dict):
    t = typed.strip()
    if not t:
        return None
    return students.get(t.casefold())


def get_suggestions(typed: str, names: list[str], n=6):
    t = typed.strip()
    if not t:
        return []
    name_map = {nm.casefold(): nm for nm in names}
    keys = list(name_map.keys())
    close = difflib.get_close_matches(t.casefold(), keys, n=n, cutoff=0.70)
    return [name_map[c] for c in close]


# -------------------------
# Present + Late marking
# -------------------------
def mark_present(wb, name: str, status: str):
    now_mt = datetime.now(MT_TZ)
    late_cutoff = now_mt.replace(hour=LATE_HOUR, minute=LATE_MINUTE, second=0, microsecond=0)

    ws = get_or_create_attendance_sheet(wb)
    date_label = today_col_label()
    col = ensure_date_column(ws, date_label)
    row = find_or_create_student_row(ws, name)

    cell = ws.cell(row=row, column=col)

    # Prevent duplicates for today
    if cell.value is not None and str(cell.value).strip() != "":
        return date_label, False, f"Already signed in for {date_label}."

    # Store time in the matrix cell
    cell.value = now_mt.time()
    cell.number_format = "h:mm AM/PM"

    if now_mt > late_cutoff:
        cell.fill = LATE_FILL

    autosize_columns(ws)

    # Log P
    log_ws = get_or_create_log_sheet(wb)
    log_ws.append([now_mt.replace(tzinfo=None), name, status, date_label, "P"])
    log_ws.cell(row=log_ws.max_row, column=1).number_format = "yyyy-mm-dd h:mm AM/PM"
    autosize_columns(log_ws)

    wb.save(WORKBOOK_PATH)
    time_str = now_mt.strftime("%I:%M %p").lstrip("0")
    return date_label, True, f"Signed in ({status}) at {time_str} MT on {date_label}."


# -------------------------
# Finalize day: fill A + produce full P/A log
# -------------------------
def finalize_today(wb):
    """
    Finalize the session for today:
    - Ensure all Students exist as rows in Attendance
    - Ensure today's date column exists
    - Fill "A" for blanks in today's date column
    - Ensure Attendance Log has one row per student for today:
        - P for present
        - A for absent
      Timestamp blank for A; for P, timestamp is taken from the matrix time if missing.
    """
    students, _ = load_students(wb)
    date_label = today_col_label()

    att_ws = get_or_create_attendance_sheet(wb)
    col = ensure_date_column(att_ws, date_label)

    # Map existing attendance rows
    row_map = {}
    for r in range(2, att_ws.max_row + 1):
        v = att_ws.cell(row=r, column=1).value
        if v:
            row_map[str(v).strip().casefold()] = r

    # Ensure every student has a row
    for key, (official_name, _status) in students.items():
        if key not in row_map:
            row_map[key] = find_or_create_student_row(att_ws, official_name)

    # Existing log entries for (date, name_key)
    log_ws = get_or_create_log_sheet(wb)
    existing = set()
    for row in log_ws.iter_rows(min_row=2, values_only=True):
        if not row:
            continue
        nm = str(row[1]).strip().casefold() if row[1] else ""
        dt = str(row[3]).strip() if len(row) > 3 and row[3] else ""
        if nm and dt:
            existing.add((dt, nm))

    # Fill A + write missing log rows
    for key, (official_name, status) in students.items():
        r = row_map[key]
        cell = att_ws.cell(row=r, column=col)
        val = cell.value

        # Determine P/A
        if val is not None and str(val).strip() != "" and str(val).strip().upper() != "A":
            pa = "P"
            # If we need to create a missing P log row, build timestamp using the time in the cell
            ts = None
            if hasattr(val, "hour"):  # time object
                now_mt = datetime.now(MT_TZ)
                ts = now_mt.replace(
                    hour=val.hour,
                    minute=val.minute,
                    second=getattr(val, "second", 0),
                    microsecond=0
                ).replace(tzinfo=None)
        else:
            # Absent: mark A in the matrix
            cell.value = "A"
            pa = "A"
            ts = ""  # blank timestamp for A

        # Write log if missing
        if (date_label, key) not in existing:
            log_ws.append([ts, official_name, status, date_label, pa])
            if ts not in ("", None):
                log_ws.cell(row=log_ws.max_row, column=1).number_format = "yyyy-mm-dd h:mm AM/PM"

    autosize_columns(att_ws)
    autosize_columns(log_ws)
    wb.save(WORKBOOK_PATH)
    return date_label


# -------------------------
# UI
# -------------------------
class AttendanceApp:
    def __init__(self, root):
        self.root = root
        self.root.title("BUI Attendance")
        self.root.configure(bg="white")

        # Kiosk full screen (ESC to exit)
        self.KIOSK_MODE = True
        if self.KIOSK_MODE:
            self.root.attributes("-fullscreen", True)
            self.root.bind("<Escape>", lambda e: self.root.attributes("-fullscreen", False))

        self.reload_registry()

        tk.Label(root, text="BUI Attendance", font=("Arial", 34, "bold"), bg="white").pack(pady=(40, 10))
        tk.Label(root, text="Type your full name and press Enter", font=("Arial", 18), bg="white").pack(pady=(0, 20))

        self.name_var = tk.StringVar()
        self.entry = tk.Entry(root, textvariable=self.name_var, font=("Arial", 26), width=32, justify="center")
        self.entry.pack(pady=10)
        self.entry.focus()

        self.status = tk.Label(root, text="", font=("Arial", 16), bg="white")
        self.status.pack(pady=(10, 10))

        tk.Label(root, text="Suggestions (click one if yours is here):", font=("Arial", 14), bg="white").pack(pady=(10, 5))
        self.sugg_list = tk.Listbox(root, font=("Arial", 16), height=6, width=34)
        self.sugg_list.pack(pady=(0, 15))
        self.sugg_list.bind("<<ListboxSelect>>", self.on_pick_suggestion)

        btn_frame = tk.Frame(root, bg="white")
        btn_frame.pack(pady=10)

        tk.Button(btn_frame, text="Submit", font=("Arial", 18), command=self.on_submit).grid(row=0, column=0, padx=10)
        tk.Button(btn_frame, text="Add New Student (Unregistered)", font=("Arial", 18), command=self.open_add_student).grid(row=0, column=1, padx=10)
        tk.Button(btn_frame, text="Refresh Registry", font=("Arial", 18), command=self.reload_registry_ui).grid(row=0, column=2, padx=10)

        tk.Button(
            btn_frame,
            text="Finalize Today (Fill A + Log)",
            font=("Arial", 18),
            command=self.on_finalize
        ).grid(row=1, column=0, columnspan=3, pady=10)

        self.root.bind("<Return>", self.on_submit)
        self.entry.bind("<KeyRelease>", lambda e: self.refresh_suggestions())

        tk.Label(
            root,
            text=f"Workbook saves to: {WORKBOOK_PATH}",
            font=("Arial", 12),
            bg="white",
            fg="#666"
        ).pack(side="bottom", pady=20)

    def reload_registry(self):
        ensure_data_dir()
        self.wb = get_or_create_workbook(WORKBOOK_PATH)
        get_or_create_students_sheet(self.wb)
        get_or_create_attendance_sheet(self.wb)
        get_or_create_log_sheet(self.wb)
        self.students, self.names = load_students(self.wb)

    def reload_registry_ui(self):
        self.reload_registry()
        self.set_status("Registry refreshed.", ok=True)
        self.refresh_suggestions()

    def set_status(self, msg, ok=True):
        self.status.config(text=msg, fg=("green" if ok else "red"))

    def refresh_suggestions(self):
        self.sugg_list.delete(0, tk.END)
        typed = self.name_var.get()
        for s in get_suggestions(typed, self.names):
            self.sugg_list.insert(tk.END, s)

    def on_pick_suggestion(self, event=None):
        sel = self.sugg_list.curselection()
        if not sel:
            return
        picked = self.sugg_list.get(sel[0])
        self.name_var.set(picked)
        self.entry.icursor(tk.END)
        self.entry.focus()
        self.set_status("Selected suggested name. Press Enter to sign in.", ok=True)

    def on_submit(self, event=None):
        typed = self.name_var.get().strip()
        if not typed:
            self.set_status("Please enter your name.", ok=False)
            return

        self.reload_registry()

        matched = canonical_match(typed, self.students)
        if matched:
            official_name, status = matched
            date_label, did_mark, msg = mark_present(self.wb, official_name, status)
            self.set_status(f"{official_name}: {msg}", ok=did_mark)

            self.name_var.set("")
            self.sugg_list.delete(0, tk.END)
            self.entry.focus()
            return

        self.refresh_suggestions()
        if self.sugg_list.size() > 0:
            self.set_status("Name not recognized. Pick a suggestion or re-type.", ok=False)
        else:
            self.set_status("Name not recognized. Ask an admin to add you.", ok=False)

    def on_finalize(self):
        self.reload_registry()
        date_label = finalize_today(self.wb)
        self.set_status(f"Finalized {date_label}: Absences marked + Log created.", ok=True)

    def open_add_student(self):
        win = tk.Toplevel(self.root)
        win.title("Add New Student (Unregistered)")
        win.configure(bg="white")
        win.geometry("650x280")

        tk.Label(win, text="Add New Student (Unregistered)", font=("Arial", 18, "bold"), bg="white").pack(pady=(15, 10))
        tk.Label(
            win,
            text="Only use this if the student is NOT in the official system.\nThey will be marked as Unregistered in Students.",
            font=("Arial", 12),
            bg="white"
        ).pack(pady=(0, 10))

        new_var = tk.StringVar()
        e = tk.Entry(win, textvariable=new_var, font=("Arial", 18), width=32, justify="center")
        e.pack(pady=10)
        e.focus()

        msg = tk.Label(win, text="", font=("Arial", 12), bg="white")
        msg.pack(pady=5)

        def do_add():
            name = new_var.get().strip()
            if not name:
                msg.config(text="Name cannot be empty.", fg="red")
                return

            if not messagebox.askyesno("Confirm", f"Add '{name}' as Unregistered?"):
                return

            self.reload_registry()
            ok, m = add_student_as_unregistered(self.wb, name)
            msg.config(text=m, fg=("green" if ok else "red"))

            if ok:
                self.reload_registry_ui()
                self.name_var.set(name)
                self.refresh_suggestions()

        tk.Button(win, text="Add as Unregistered", font=("Arial", 14), command=do_add).pack(pady=10)
        tk.Button(win, text="Close", font=("Arial", 12), command=win.destroy).pack(pady=5)
        win.bind("<Return>", lambda e: do_add())


def main():
    root = tk.Tk()
    AttendanceApp(root)
    root.mainloop()


if __name__ == "__main__":
    main()
