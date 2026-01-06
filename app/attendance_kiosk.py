"""
BUI Attendance Kiosk (Tkinter + Excel)

What this does:
- Students type their name into a simple kiosk UI.
- Attendance logs into an Excel workbook.
- Each day gets its own sheet named YYYY-MM-DD.
- A master sheet "Students" stores:
    Name | OfficialStatus
  where OfficialStatus is "Registered" or "Unregistered".
- If a typed name EXACTLY matches a student in Students (case-insensitive):
    -> attendance is logged with their status.
- If name is NOT in Students:
    -> attendance is blocked
    -> suggestions appear (close spelling matches)
    -> admin can add them, but they are ALWAYS added as "Unregistered"
- No duplicates per day: the same student cannot be logged twice on the same date sheet.

Best-practice notes:
- Keep data files in /data, code in /app.
- Use a virtual environment (.venv) and install openpyxl.
- Avoid hardcoding student lists in code; keep them in the Students sheet.
"""

import os
from datetime import datetime
import difflib
import tkinter as tk
from tkinter import messagebox
from openpyxl import Workbook, load_workbook
from openpyxl.utils import get_column_letter


# -------------------------
# Paths / constants
# -------------------------
PROJECT_ROOT = os.path.dirname(os.path.dirname(__file__))  # .../BUI_Attendance_Kiosk
DATA_DIR = os.path.join(PROJECT_ROOT, "data")
WORKBOOK_PATH = os.path.join(DATA_DIR, "BUI_Attendance.xlsx")

DATE_FORMAT = "%Y-%m-%d"      # daily sheet name like 2026-01-06
STUDENTS_SHEET = "Students"   # master list
STATUS_REGISTERED = "Registered"
STATUS_UNREGISTERED = "Unregistered"


# -------------------------
# Excel helpers
# -------------------------
def ensure_data_dir():
    """Create required folders automatically (best practice)."""
    os.makedirs(DATA_DIR, exist_ok=True)


def autosize_columns(ws):
    """Autosize columns for readability (simple approach)."""
    for col in range(1, ws.max_column + 1):
        max_len = 0
        col_letter = get_column_letter(col)
        for cell in ws[col_letter]:
            if cell.value is not None:
                max_len = max(max_len, len(str(cell.value)))
        ws.column_dimensions[col_letter].width = min(max_len + 2, 45)


def get_or_create_workbook(path):
    """Open an existing workbook or create a new one."""
    if os.path.exists(path):
        return load_workbook(path)

    wb = Workbook()
    wb.save(path)
    return load_workbook(path)


def get_or_create_students_sheet(wb):
    """
    Ensure the master Students sheet exists with headers:
      Name | OfficialStatus

    Also upgrades old versions that used "Registered Student Name".
    """
    if STUDENTS_SHEET in wb.sheetnames:
        ws = wb[STUDENTS_SHEET]
        # Upgrade old header if needed
        a1 = ws.cell(row=1, column=1).value
        b1 = ws.cell(row=1, column=2).value if ws.max_column >= 2 else None

        if a1 == "Registered Student Name" or (a1 == "Name" and b1 is None):
            ws.cell(row=1, column=1, value="Name")
            ws.cell(row=1, column=2, value="OfficialStatus")
            # Default existing rows to Registered if status missing
            for r in range(2, ws.max_row + 1):
                name_val = ws.cell(row=r, column=1).value
                status_val = ws.cell(row=r, column=2).value
                if name_val and not status_val:
                    ws.cell(row=r, column=2, value=STATUS_REGISTERED)
            autosize_columns(ws)
            wb.save(WORKBOOK_PATH)

        return ws

    ws = wb.create_sheet(STUDENTS_SHEET)
    ws.append(["Name", "OfficialStatus"])
    autosize_columns(ws)
    wb.save(WORKBOOK_PATH)
    return ws


def load_students(wb):
    """
    Load students from Students sheet.

    Returns:
      students: dict casefold(name) -> (OfficialName, OfficialStatus)
      names: list of OfficialName (for suggestions / display)
    """
    ws = get_or_create_students_sheet(wb)
    students = {}
    names = []

    for row in ws.iter_rows(min_row=2, values_only=True):
        if not row or not row[0]:
            continue
        name = str(row[0]).strip()
        status = str(row[1]).strip() if len(row) > 1 and row[1] else STATUS_UNREGISTERED

        key = name.casefold()
        # Deduplicate by case-insensitive key (keep first occurrence)
        if key not in students:
            students[key] = (name, status)
            names.append(name)

    return students, names


def add_student_as_unregistered(wb, name: str):
    """Add a new student to Students sheet as Unregistered (always)."""
    name = name.strip()
    if not name:
        return False, "Name cannot be empty."

    students, _ = load_students(wb)
    if name.casefold() in students:
        return False, "That student already exists in Students."

    ws = get_or_create_students_sheet(wb)
    ws.append([name, STATUS_UNREGISTERED])
    autosize_columns(ws)
    wb.save(WORKBOOK_PATH)
    return True, f"Added as Unregistered: {name}"


def get_or_create_daily_sheet(wb):
    """Create today's attendance sheet named YYYY-MM-DD with headers."""
    sheet_name = datetime.now().strftime(DATE_FORMAT)
    if sheet_name not in wb.sheetnames:
        ws = wb.create_sheet(sheet_name)
        ws.append(["Time", "Name", "OfficialStatus"])
        autosize_columns(ws)
        wb.save(WORKBOOK_PATH)
    return wb[sheet_name], sheet_name


def already_signed_in_today(ws, name: str) -> bool:
    """True if 'name' already exists on today's sheet (case-insensitive)."""
    target = name.strip().casefold()
    if not target:
        return False

    # Columns: Time | Name | OfficialStatus
    for row in ws.iter_rows(min_row=2, values_only=True):
        if not row or len(row) < 2 or not row[1]:
            continue
        existing_name = str(row[1]).strip().casefold()
        if existing_name == target:
            return True
    return False


def log_attendance(wb, name: str, status: str):
    """
    Log attendance for today.
    Returns: (sheet_name, did_log_bool)
    """
    ws, sheet_name = get_or_create_daily_sheet(wb)

    if already_signed_in_today(ws, name):
        return sheet_name, False

    time_str = datetime.now().strftime("%H:%M:%S")
    ws.append([time_str, name, status])
    autosize_columns(ws)
    wb.save(WORKBOOK_PATH)
    return sheet_name, True


# -------------------------
# Matching / validation
# -------------------------
def canonical_match(typed: str, students: dict):
    """
    Return (OfficialName, OfficialStatus) if typed matches a student (case-insensitive), else None.
    students: dict casefold(name) -> (OfficialName, OfficialStatus)
    """
    t = typed.strip()
    if not t:
        return None
    key = t.casefold()
    return students.get(key)


def get_suggestions(typed: str, names: list[str], n=6):
    """Return close-match suggestions for spelling mistakes."""
    t = typed.strip()
    if not t:
        return []

    # Build casefold map so difflib matches are case-insensitive but we display official spellings
    name_map = {nm.casefold(): nm for nm in names}
    keys = list(name_map.keys())

    close = difflib.get_close_matches(t.casefold(), keys, n=n, cutoff=0.70)
    return [name_map[c] for c in close]


# -------------------------
# UI App
# -------------------------
class AttendanceApp:
    def __init__(self, root):
        self.root = root
        self.root.title("BUI Attendance")
        self.root.configure(bg="white")

        # Kiosk full screen (ESC to exit)
        self.KIOSK_MODE = True
        if self.KIOSK_MODE:
            self.root.attributes("-fullscreen", True)
            self.root.bind("<Escape>", lambda e: self.root.attributes("-fullscreen", False))

        self.reload_registry()

        # Title
        tk.Label(root, text="BUI Attendance", font=("Arial", 34, "bold"), bg="white").pack(pady=(40, 10))
        tk.Label(root, text="Type your full name and press Enter", font=("Arial", 18), bg="white").pack(pady=(0, 20))

        # Entry
        self.name_var = tk.StringVar()
        self.entry = tk.Entry(root, textvariable=self.name_var, font=("Arial", 26), width=32, justify="center")
        self.entry.pack(pady=10)
        self.entry.focus()

        # Status
        self.status = tk.Label(root, text="", font=("Arial", 16), bg="white")
        self.status.pack(pady=(10, 10))

        # Suggestions
        tk.Label(root, text="Suggestions (click one if yours is here):", font=("Arial", 14), bg="white").pack(pady=(10, 5))
        self.sugg_list = tk.Listbox(root, font=("Arial", 16), height=6, width=34)
        self.sugg_list.pack(pady=(0, 15))
        self.sugg_list.bind("<<ListboxSelect>>", self.on_pick_suggestion)

        # Buttons
        btn_frame = tk.Frame(root, bg="white")
        btn_frame.pack(pady=10)

        tk.Button(btn_frame, text="Submit", font=("Arial", 18), command=self.on_submit).grid(row=0, column=0, padx=10)
        tk.Button(btn_frame, text="Add New Student (Unregistered)", font=("Arial", 18), command=self.open_add_student).grid(row=0, column=1, padx=10)
        tk.Button(btn_frame, text="Refresh Registry", font=("Arial", 18), command=self.reload_registry_ui).grid(row=0, column=2, padx=10)

        # Bindings
        self.root.bind("<Return>", self.on_submit)
        self.entry.bind("<KeyRelease>", lambda e: self.refresh_suggestions())

        # Footer
        tk.Label(
            root,
            text=f"Workbook saves to: {WORKBOOK_PATH}",
            font=("Arial", 12),
            bg="white",
            fg="#666"
        ).pack(side="bottom", pady=20)

    # ---- Registry management ----
    def reload_registry(self):
        ensure_data_dir()
        self.wb = get_or_create_workbook(WORKBOOK_PATH)
        get_or_create_students_sheet(self.wb)
        self.students, self.names = load_students(self.wb)

    def reload_registry_ui(self):
        self.reload_registry()
        self.set_status("Registry refreshed.", ok=True)
        self.refresh_suggestions()

    # ---- UI helpers ----
    def set_status(self, msg, ok=True):
        self.status.config(text=msg, fg=("green" if ok else "red"))

    def refresh_suggestions(self):
        self.sugg_list.delete(0, tk.END)
        typed = self.name_var.get()
        for s in get_suggestions(typed, self.names):
            self.sugg_list.insert(tk.END, s)

    def on_pick_suggestion(self, event=None):
        sel = self.sugg_list.curselection()
        if not sel:
            return
        picked = self.sugg_list.get(sel[0])
        self.name_var.set(picked)
        self.entry.icursor(tk.END)
        self.entry.focus()
        self.set_status("Selected suggested name. Press Enter to sign in.", ok=True)

    # ---- Attendance flow ----
    def on_submit(self, event=None):
        typed = self.name_var.get().strip()
        if not typed:
            self.set_status("Please enter your name.", ok=False)
            return

        # Re-load to keep registry current if someone edited Excel
        self.reload_registry()

        matched = canonical_match(typed, self.students)
        if matched:
            official_name, status = matched
            sheet_name, did_log = log_attendance(self.wb, official_name, status)

            if did_log:
                self.set_status(f"Signed in: {official_name} ({status})", ok=True)
            else:
                self.set_status(f"Already signed in today: {official_name}", ok=False)

            self.name_var.set("")
            self.sugg_list.delete(0, tk.END)
            self.entry.focus()
            return

        # Block attendance if not in Students
        self.refresh_suggestions()
        if self.sugg_list.size() > 0:
            self.set_status("Name not recognized. Pick a suggestion or re-type.", ok=False)
        else:
            self.set_status("Name not recognized. Ask an admin to add you.", ok=False)

    # ---- Add student ----
    def open_add_student(self):
        win = tk.Toplevel(self.root)
        win.title("Add New Student (Unregistered)")
        win.configure(bg="white")
        win.geometry("600x260")

        tk.Label(win, text="Add New Student (Unregistered)", font=("Arial", 18, "bold"), bg="white").pack(pady=(15, 10))
        tk.Label(
            win,
            text="Only use this if the student is NOT in the official system.\nThey will be marked as Unregistered.",
            font=("Arial", 12),
            bg="white"
        ).pack(pady=(0, 10))

        new_var = tk.StringVar()
        e = tk.Entry(win, textvariable=new_var, font=("Arial", 18), width=32, justify="center")
        e.pack(pady=10)
        e.focus()

        msg = tk.Label(win, text="", font=("Arial", 12), bg="white")
        msg.pack(pady=5)

        def do_add():
            name = new_var.get().strip()
            if not name:
                msg.config(text="Name cannot be empty.", fg="red")
                return

            if not messagebox.askyesno("Confirm", f"Add '{name}' as Unregistered?"):
                return

            self.reload_registry()
            ok, m = add_student_as_unregistered(self.wb, name)
            msg.config(text=m, fg=("green" if ok else "red"))

            if ok:
                self.reload_registry_ui()
                self.name_var.set(name)
                self.refresh_suggestions()

        tk.Button(win, text="Add as Unregistered", font=("Arial", 14), command=do_add).pack(pady=10)
        tk.Button(win, text="Close", font=("Arial", 12), command=win.destroy).pack(pady=5)
        win.bind("<Return>", lambda e: do_add())


def main():
    root = tk.Tk()
    AttendanceApp(root)
    root.mainloop()


if __name__ == "__main__":
    main()
